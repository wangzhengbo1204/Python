#-*- coding:utf-8 -*-
'''
Created on 2012-7-11

@author: GFTOwenWang
'''
from openpyxl.reader.excel import load_workbook
from pyExcelerator import parse_xls


def parse():
    filefullname = r'D:\GFT\Test\test_excel.xlsx'
    wb = load_workbook(filename=filefullname)
    ws = wb.get_sheet_by_name("Sheet1")
    ws.cell("A3")
    ws.cell("A4")
    ws.cell("A5")
    
def parse2():
    filefullname = r'D:\GFT\Test\test_excel.xls'
    sheet_values = parse_xls(filefullname)
    print sheet_values
    

if __name__ == '__main__':
    parse()